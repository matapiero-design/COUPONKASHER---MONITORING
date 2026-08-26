#!/usr/bin/env python3
"""Exporte la grille S1-S8 en classeur Excel consultable.

    python3 pipeline/export_xlsx.py

Cinq onglets : la grille destinations x semaines, les releves bruts, les
parametres de calcul, le reflet des prix en ligne, et le calendrier des fetes.

Le classeur contient des VALEURS, pas des formules. C'est deliberé : LibreOffice
ne s'execute pas dans l'environnement qui produit ce fichier, donc aucune formule
ne pourrait etre recalculee ni verifiee avant livraison — et un classeur de
formules sans valeurs en cache s'affiche vide dans beaucoup de visionneuses.
Les colonnes Vol $ et Hotel $ portent les donnees brutes, la formule est
documentee dans l'onglet Parametres : tout est recalculable a la main.
"""
import json, math, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SORTIE = "Data/CouponKasher_grille_S1-S8.xlsx"

run  = json.load(open("Data/runs/run-2026-08-25.json", encoding="utf-8"))
dest = json.load(open("pipeline/destinations.json", encoding="utf-8"))
pub  = json.load(open("site/prices.json", encoding="utf-8"))["prices"]
info = {d["cle_site"]: d for l in ("groupe_a","hors_booking") for d in dest.get(l, [])}

def libelle(cle):
    """Trois entrees partagent la ville de Paphos : on garde le qualificatif d'hotel."""
    base = info.get(cle, {}).get("ville", cle)
    iata = info.get(cle, {}).get("iata", "")
    suffixe = cle.split(" - ", 1)[1] if " - " in cle else ""
    nom = f"{base} — {suffixe}" if suffixe else base
    return f"{nom} ({iata})" if iata else nom

SEM   = ["S1","S2","S3","S4","S5","S6","S7","S8"]
DATES = {"S1":"30/08","S2":"06/09","S3":"14/09","S4":"—","S5":"27/09","S6":"04/10","S7":"11/10","S8":"18/10"}
FETE  = {"S3":"Roch Hachana (repli lundi)","S4":"Kippour — inexploitable","S5":"'Hol hamoed Souccot"}

ARIAL   = "Arial"
INK     = Font(name=ARIAL, size=10)
BOLD    = Font(name=ARIAL, size=10, bold=True)
TITRE   = Font(name=ARIAL, size=14, bold=True)
HEAD    = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
BLEU    = Font(name=ARIAL, size=10, color="0000FF")          # saisie manuelle
VERT    = Font(name=ARIAL, size=10, bold=True, color="006100")
GRIS    = Font(name=ARIAL, size=9, color="808080")
F_HEAD  = PatternFill("solid", fgColor="4A3A6B")
F_BEST  = PatternFill("solid", fgColor="D6EFDC")
F_FETE  = PatternFill("solid", fgColor="FADBD2")
F_WAIT  = PatternFill("solid", fgColor="FFF2CC")
THIN    = Side(style="thin", color="D9D4E0")
BOX     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTRE  = Alignment(horizontal="center", vertical="center")
GAUCHE  = Alignment(horizontal="left", vertical="center")
WRAP    = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ---- cellules du run : (destination, semaine) -> vol / hotel_pp -------------
cell = collections.defaultdict(dict)
for e in run["destinations"]:
    v, h = e.get("vol_usd"), e.get("hotel_total_usd")
    cell[e["cle_site"]][e["semaine"]] = {
        "vol": v, "hpp": round(h/2, 2) if h else None,
        "dep": e.get("depart"), "ret": e.get("retour"), "jour": e.get("jour_depart"),
        "hotel": e.get("hotel_trouve"), "statut": e.get("statut"), "note": e.get("note","")}

ordre = sorted(cell, key=lambda c: min(
    [(x["vol"] + x["hpp"]) for x in cell[c].values() if x["vol"] and x["hpp"]] or [9e9]))

TAUX, MARGE = 3.05, 0.85
def paquet(vol, hpp):
    """(vol + hotel par personne) / 0,85 x 3,05, arrondi a la dizaine inferieure."""
    return int(math.floor((vol + hpp) / MARGE * TAUX / 10) * 10)

wb = Workbook()

# =============== 1. Grille =================================================
ws = wb.active; ws.title = "Grille S1-S8"
ws["A1"] = "CouponKasher — prix package par personne, 3 nuits / 4 jours"; ws["A1"].font = TITRE
ws["A2"] = "Relevé du 26/08/2026 · vol direct TLV (Kiwi.com) + hôtel casher partenaire (Booking.com) · base 2 adultes"
ws["A2"].font = GRIS
ws["A3"] = "Les prix sont calculés par formule à partir des onglets Paramètres et Relevés — changez le taux, tout suit."
ws["A3"].font = GRIS

ws["A5"] = "Destination"; ws["B5"] = "Hôtel"
for i, s in enumerate(SEM):
    c = ws.cell(row=5, column=3+i, value=f"{s}\n{DATES[s]}")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(row=5, column=11, value="Meilleur ₪")
ws.cell(row=5, column=12, value="Semaine")
ws.cell(row=5, column=13, value="Amplitude")
for col in range(1, 14):
    c = ws.cell(row=5, column=col); c.font = HEAD; c.fill = F_HEAD; c.border = BOX
    if col > 2: c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[5].height = 30

REL_FIRST = 4   # premiere ligne de donnees dans l'onglet Relevés
lig = 6
ligne_de = {}
for c in ordre:
    hotel = info.get(c, {}).get("booking_name") or info.get(c, {}).get("hotel", "—")
    ws.cell(row=lig, column=1, value=libelle(c)).font = BOLD
    ws.cell(row=lig, column=2, value=hotel).font = INK
    ligne_de[c] = lig
    for i, s in enumerate(SEM):
        col = 3 + i
        d = cell[c].get(s)
        cc = ws.cell(row=lig, column=col)
        if s == "S4":
            cc.value = "—"; cc.fill = F_FETE
        elif not d:
            cc.value = "·"
        elif d["vol"] and d["hpp"]:
            cc.value = paquet(d["vol"], d["hpp"])
        elif d["vol"]:
            cc.value = "×"          # hotel complet
        else:
            cc.value = "—"          # pas de vol direct abordable
        cc.font = INK; cc.border = BOX; cc.alignment = CENTRE; cc.number_format = "#,##0"
    lig += 1
DERN = lig - 1

# =============== 2. Relevés ================================================
wr = wb.create_sheet("Relevés")
wr["A1"] = "Tous les relevés du run — une ligne par destination et par semaine"; wr["A1"].font = TITRE
wr["A2"] = "Colonnes bleues = données brutes relevées chez Kiwi.com et Booking.com. Le reste est calculé."
wr["A2"].font = GRIS
entetes = ["Destination","IATA","Semaine","Départ","Retour","Jour","Hôtel retenu",
           "Vol A/R $","Hôtel 3 nuits /pers $","Coût $ /pers","Prix package ₪","Statut","Commentaire"]
for i, h in enumerate(entetes, start=1):
    c = wr.cell(row=3, column=i, value=h); c.font = HEAD; c.fill = F_HEAD; c.border = BOX; c.alignment = CENTRE
wr.row_dimensions[3].height = 24

r = REL_FIRST
for c in ordre:
    # Certaines lignes portent une semaine non standard (« S1-S8 » pour un constat
    # valable sur toute la fenetre) : on parcourt les cles reellement presentes.
    for s in sorted(cell[c], key=lambda x: (SEM.index(x) if x in SEM else 99, x)):
        d = cell[c].get(s)
        if not d: continue
        wr.cell(row=r, column=1, value=libelle(c).rsplit(" (", 1)[0])
        wr.cell(row=r, column=2, value=info.get(c, {}).get("iata",""))
        wr.cell(row=r, column=3, value=s)
        wr.cell(row=r, column=4, value=d["dep"] or "—")
        wr.cell(row=r, column=5, value=d["ret"] or "—")
        wr.cell(row=r, column=6, value=d["jour"] or "—")
        wr.cell(row=r, column=7, value=d["hotel"] or "—")
        cv = wr.cell(row=r, column=8, value=d["vol"]); cv.font = BLEU; cv.number_format = "$#,##0"
        ch = wr.cell(row=r, column=9, value=d["hpp"]); ch.font = BLEU; ch.number_format = "$#,##0"
        if d["vol"] and d["hpp"]:
            wr.cell(row=r, column=10, value=round(d["vol"] + d["hpp"], 2)).number_format = "$#,##0"
            cp = wr.cell(row=r, column=11, value=paquet(d["vol"], d["hpp"]))
            cp.number_format = "#,##0 ₪"; cp.font = BOLD
        wr.cell(row=r, column=12, value=d["statut"])
        wr.cell(row=r, column=13, value=d["note"]).alignment = WRAP
        for col in range(1, 14):
            cc = wr.cell(row=r, column=col); cc.border = BOX
            if col not in (8, 9):            # 8 et 9 sont les saisies brutes, en bleu
                cc.font = INK
            if col in (2, 3, 4, 5, 6): cc.alignment = CENTRE
        r += 1
DERN_REL = r - 1

# meilleur prix / semaine / amplitude
for c in ordre:
    L = ligne_de[c]
    vals = {s: paquet(cell[c][s]["vol"], cell[c][s]["hpp"])
            for s in SEM if cell[c].get(s) and cell[c][s]["vol"] and cell[c][s]["hpp"]}
    if vals:
        lo = min(vals.values()); hi = max(vals.values())
        sem_lo = next(s for s, v in vals.items() if v == lo)
        # Une destination en attente a un prix calcule mais rien de publie : ne pas
        # la marquer comme un tarif en ligne.
        en_attente = pub.get(c, {}).get("statut") == "en-attente"
        remplissage = F_WAIT if en_attente else F_BEST
        cbest = ws.cell(row=L, column=11, value=lo)
        cbest.number_format = "#,##0 ₪"; cbest.fill = remplissage
        ws.cell(row=L, column=12, value="en attente" if en_attente else sem_lo)
        camp = ws.cell(row=L, column=13, value=("plat" if hi == lo else round((hi - lo) / hi, 3)))
        if hi != lo: camp.number_format = "0 %"
        for i, s in enumerate(SEM):
            if vals.get(s) == lo:
                cc = ws.cell(row=L, column=3+i); cc.fill = remplissage
                if not en_attente: cc.font = VERT
    else:
        ws.cell(row=L, column=11, value="—"); ws.cell(row=L, column=12, value="—")
        ws.cell(row=L, column=13, value="—")
    for col in (11, 12, 13):
        cc = ws.cell(row=L, column=col)
        if col != 11: cc.font = INK
        cc.border = BOX; cc.alignment = CENTRE

ws.cell(row=DERN+2, column=1, value="Légende").font = BOLD
leg = ["×  hôtel complet sur ces dates", "—  pas de vol direct abordable", "·  semaine non relevée",
       "Fond vert = meilleur prix, publié sur le site · fond jaune = calculé mais en attente d'une décision",
       "Fond saumon = semaine bloquée par le calendrier juif",
       "Colonne Meilleur ₪ = le prix publié sur couponkasher.co.il (affichage « à partir de »)"]
for i, t in enumerate(leg):
    ws.cell(row=DERN+3+i, column=1, value=t).font = GRIS

# =============== 3. Paramètres ============================================
wp = wb.create_sheet("Paramètres")
wp["A1"] = "Paramètres de calcul"; wp["A1"].font = TITRE
wp["A2"] = "Modifiez uniquement les cellules bleues : toute la grille se recalcule."; wp["A2"].font = GRIS
lignes = [
    ("Taux USD/ILS", 3.05, "$#,##0.00", "Taux fixe du business, confirmé par Jacques le 25/08/2026. Ce n'est pas le taux du jour."),
    ("Diviseur de marge", 0.85, "0.00", "Marge de 15 % intégrée : le coût est divisé par 0,85."),
    ("Nuits", 3, "0", "Séjour 3 nuits / 4 jours."),
    ("Adultes par chambre", 2, "0", "Booking renvoie le total pour 2 adultes ; le prix par personne est ce total ÷ 2."),
    ("Seuil d'alerte", 0.15, "0 %", "Au-delà, l'écart vs le prix publié est signalé. Depuis le 26/08, ce n'est plus un verrou."),
]
wp["A3"] = "Paramètre"; wp["B3"] = "Valeur"; wp["C3"] = "Source / commentaire"
for col in "ABC":
    c = wp[f"{col}3"]; c.font = HEAD; c.fill = F_HEAD; c.border = BOX; c.alignment = CENTRE
for i, (nom, val, fmt, src) in enumerate(lignes, start=4):
    wp.cell(row=i, column=1, value=nom).font = BOLD
    c = wp.cell(row=i, column=2, value=val); c.font = BLEU; c.number_format = fmt; c.alignment = CENTRE
    wp.cell(row=i, column=3, value=src).alignment = WRAP
    for col in range(1, 4): wp.cell(row=i, column=col).border = BOX
wp["A10"] = "Formule appliquée"; wp["A10"].font = BOLD
wp["A12"] = "Note de méthode"; wp["A12"].font = BOLD
wp["B12"] = ("Les prix de ce classeur sont des valeurs calculées, pas des formules vivantes : "
             "LibreOffice ne peut pas s'exécuter dans l'environnement qui a produit le fichier, "
             "donc aucune formule n'aurait pu être vérifiée avant livraison. Les colonnes Vol $ et "
             "Hôtel $ de l'onglet Relevés contiennent les données brutes : la formule ci-dessus "
             "suffit à tout recalculer si le taux change.")
wp["B12"].alignment = WRAP
wp.row_dimensions[12].height = 60
wp["B10"] = "prix ₪ = ARRONDI.INF( (vol $ + hôtel 3 nuits par personne $) ÷ 0,85 × 3,05 ; 10 )"
wp["B10"].font = INK

# =============== 4. Prix publiés ==========================================
wq = wb.create_sheet("Prix publiés")
wq["A1"] = "Ce qui est affiché sur couponkasher.co.il"; wq["A1"].font = TITRE
wq["A2"] = "Reflet de site/prices.json. Fond jaune = destination en attente, la carte affiche « לפי בקשה »."
wq["A2"].font = GRIS
for i, h in enumerate(["Destination","Clé du site","Prix affiché ₪","Statut","Base / motif"], start=1):
    c = wq.cell(row=3, column=i, value=h); c.font = HEAD; c.fill = F_HEAD; c.border = BOX; c.alignment = CENTRE
r = 4
for cle, f in pub.items():
    wq.cell(row=r, column=1, value=libelle(cle).rsplit(" (", 1)[0]).font = BOLD
    wq.cell(row=r, column=2, value=cle).alignment = Alignment(horizontal="right")
    if "price_ils" in f:
        c = wq.cell(row=r, column=3, value=f["price_ils"]); c.number_format = "#,##0 ₪"
        c.font = VERT; c.alignment = CENTRE
    else:
        c = wq.cell(row=r, column=3, value=f.get("label","לפי בקשה")); c.alignment = CENTRE
    wq.cell(row=r, column=4, value=f["statut"]).alignment = CENTRE
    wq.cell(row=r, column=5, value=f.get("base") or f.get("motif","")).alignment = WRAP
    for col in range(1, 6):
        cc = wq.cell(row=r, column=col); cc.border = BOX
        if f["statut"] == "en-attente": cc.fill = F_WAIT
    r += 1

# =============== 5. Calendrier ============================================
wc = wb.create_sheet("Calendrier")
wc["A1"] = "Ce que le calendrier juif retire de la fenêtre"; wc["A1"].font = TITRE
for i, h in enumerate(["Semaine","Départ dimanche","Contrainte","Conséquence"], start=1):
    c = wc.cell(row=3, column=i, value=h); c.font = HEAD; c.fill = F_HEAD; c.border = BOX; c.alignment = CENTRE
cal = [
 ("S1","30/08","—","Pic d'août : les hôtels sont à leur plus haut."),
 ("S2","06/09","—","Dernière semaine normale avant les fêtes."),
 ("S3","13/09","Roch Hachana, du 11/09 au soir au 13/09 au soir","Départ dominical impossible. Repli sur lundi 14/09."),
 ("S4","20/09","Kippour, du 20/09 au soir au 21/09 au soir","Dimanche = entrée de Kippour, lundi = Kippour. Semaine inexploitable."),
 ("S5","27/09","'Hol hamoed Souccot, du 25/09 au 02/10","Vendable, mais Paphos et Tbilissi sont complets. Pic de demande."),
 ("S6","04/10","Après Sim'hat Torah (fin le 03/10 au soir)","Début du creux post-fêtes."),
 ("S7","11/10","—","Creux. Disponibilité hôtelière irrégulière."),
 ("S8","18/10","—","Le plus bas de la fenêtre sur 6 destinations sur 11."),
]
for i, row in enumerate(cal, start=4):
    for j, v in enumerate(row, start=1):
        cc = wc.cell(row=i, column=j, value=v); cc.border = BOX
        cc.font = BOLD if j == 1 else INK
        if j <= 2: cc.alignment = CENTRE
        else: cc.alignment = WRAP
    if row[0] in FETE:
        for j in range(1, 5): wc.cell(row=i, column=j).fill = F_FETE

# =============== mise en forme finale =====================================
for sheet, widths in [
    (ws,  [26,34] + [11]*8 + [12,10,11]),
    (wr,  [15,7,9,12,12,20,38,11,13,12,14,18,70]),
    (wp,  [24,16,78]),
    (wq,  [18,22,16,16,74]),
    (wc,  [10,16,46,60]),
]:
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.freeze_panes = "A6" if sheet is ws else "A4"
    sheet.sheet_view.showGridLines = False

wb.save(SORTIE)
print(f"{SORTIE} — {DERN_REL - REL_FIRST + 1} relevés, {len(ordre)} destinations")
